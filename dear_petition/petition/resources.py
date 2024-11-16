import re
import typing
from wsgiref import headers
from abc import ABC, abstractmethod

from django.db.models.query import QuerySet
from django.db.models import BooleanField
from django.db import transaction
from import_export import fields, resources
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import tablib

from dear_petition.petition import models, constants
from dear_petition.petition.etl.load import create_batch_petitions

bold_font = Font(bold=True)

CLIENT_SHEET_TITLE = "Client Information"


def titlecase(string):
    string = string.replace("_", " ")
    return string.title()


def int_to_char(i: int) -> str:
    return chr(i + 65)


def parse_agency_full_address(full_address: str) -> typing.Tuple[str, str, str, str, str]:
    """Parse fields needed for agency lookup"""
    [*address_lines, city_state_zip_line] = full_address.split("\n")
    if len(address_lines) == 0:
        raise ValueError("Address must be split into multiple lines")
    adress1 = address_lines[0]
    address2 = address_lines[1] if len(address_lines) == 2 else None

    match = re.match(r"\s*([^,]+?)[,\s]\s*(\w{2})\s*([\w-]+)\s*$", city_state_zip_line)
    if not match:
        raise ValueError("City, State Zipcode format is incorrect")
    (city, state, zipcode) = match.groups()
    return (adress1, address2, city, state, zipcode)


def is_empty_row(row) -> bool:
    return all([value is None for value in row.values()])


class ExcelField(fields.Field):
    def __init__(self, **kwargs):
        self.dropdown = kwargs.pop("dropdown", None)
        super().__init__(**kwargs)


class ExcelDataset:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.ws = None

    def change_worksheet(self, sheet_name=None, sheet_index=None):
        if sheet_name:
            self.ws = self.wb[sheet_name]
        elif sheet_index is not None:
            self.ws = self.wb.worksheets[sheet_index]
        else:
            raise ValueError("You must provide either a sheet name or a sheet index.")

    def create_new_worksheet(self, title=None):
        self.wb.create_sheet(title=title)
        self.change_worksheet(sheet_name=title)

    def append(
        self,
        values,
        fields,
        is_header=False,
        num_indent=0,
        bold=False,
        color=None,
    ):
        if len(values) != len(fields):
            raise ValueError("The values and fields passed to append should match.")

        row = ["" for i in range(num_indent)] + values
        self.ws.append(row)
        row_number = self.ws.max_row

        fill = None
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Apply font and alignment to each cell in the row if specified
        for col_number, field in enumerate(fields, num_indent + 1):
            cell = self.ws.cell(row=row_number, column=col_number)
            if bold:
                cell.font = bold_font
            if fill:
                cell.fill = fill
            if not is_header and getattr(field, "dropdown", None):
                self.ws.add_data_validation(field.dropdown)
                dropdown = field.dropdown
                dropdown.add(cell)

    def append_separator(self):
        """
        Appends a blank row
        """
        self.ws.append([])

    def resize_columns(self):
        for sheet in self.wb.worksheets:
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                if max_length > 0:
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[column].width = adjusted_width

    def save(self, filename):
        """
        Save the workbook to a file.
        """
        self.wb.save(filename)


class AgencyResource(resources.ModelResource):
    county = ExcelField(attribute="county", column_name="County")
    name = ExcelField(attribute="name", column_name="Arresting Agency")
    address = ExcelField(column_name="Address")

    def init_instance(self, row=None):
        instance = super().init_instance(row)
        if is_empty_row(row):
            return instance
        instance.address1 = row["address1"]
        instance.address2 = row["address2"]
        return instance

    def get_instance(self, instance_loader, row):
        try:
            return super().get_instance(instance_loader, row)
        except models.Agency.MultipleObjectsReturned as e:
            raise models.Agency.MultipleObjectsReturned(
                f"There are multiple agencies named '{row['Arresting Agency']}' in county '{row['County']}. Ensure there is only 1."
            )

    def before_import_row(self, row, **kwargs):
        if is_empty_row(row):
            return

        # found a bug where leading newlines were causing false negative matches for existing agencies
        row["Arresting Agency"] = row["Arresting Agency"].strip()
        row["name"] = row["Arresting Agency"]
        row["County"] = row["County"].strip()
        row["county"] = row["County"].strip()

        (address1, address2, city, state, zipcode) = parse_agency_full_address(row["Address"])
        row["city"] = city.strip()
        row["state"] = state.strip()
        row["zipcode"] = zipcode.strip()
        row["address1"] = address1.strip()
        row["address2"] = address2.strip() if address2 else None

        name = row["name"]
        if re.search(models.AGENCY_SHERRIFF_OFFICE_PATTERN, name, re.IGNORECASE) or re.search(
            models.AGENCY_SHERRIFF_DEPARTMENT_PATTERN, name, re.IGNORECASE
        ):
            row["is_sheriff"] = True
        else:
            row["is_sheriff"] = False

    def skip_row(self, instance, original, row, import_validation_errors=None):
        if is_empty_row(row):
            return True
        return super().skip_row(
            instance, original, row, import_validation_errors=import_validation_errors
        )

    class Meta:
        model = models.Agency
        import_id_fields = ("name", "county")
        store_instance = True


class MultiModelResource(resources.ModelResource):
    PARENT_OBJECT_FIELD = None
    NUM_INDENT = 0

    def __init__(self):
        super().__init__()
        self.saved_instance_ids = []

    def after_save_instance(self, instance, using_transactions, dry_run):
        if not dry_run:
            self.saved_instance_ids.append(instance.id)

    def export_resource(self, obj):
        return [self.export_field(field, obj) for field in self.get_export_fields()]

    def get_export_headers(self):
        headers = [titlecase(field.column_name) for field in self.get_fields()]
        return headers

    def export_field(self, field, obj):
        field_name = self.get_field_name(field)
        dehydrate_method = field.get_dehydrate_method(field_name)

        if isinstance(self._meta.model._meta.get_field(field.attribute), BooleanField):
            field.dropdown = DataValidation(
                type="list", formula1=f'"True,False"', showDropDown=True
            )

        method = getattr(self, dehydrate_method, None)
        if method is not None:
            return method(obj)

        return field.export(obj)

    def import_field(self, field, obj, data, is_m2m=False, **kwargs):
        hydrate_method = getattr(self, f"hydrate_{field.attribute}", None)
        if hydrate_method is not None:
            hydrate_method(field, data)

        if field.attribute and field.column_name in data:
            field.save(obj, data, is_m2m, **kwargs)

    def export_excel(self, data, dataset: ExcelDataset, color=None, **kwargs):
        headers = self.get_export_headers()
        fields = self.get_export_fields()

        dataset.append(
            headers, fields, num_indent=self.NUM_INDENT, is_header=True, bold=True, color=color
        )

        if isinstance(data, QuerySet):
            for obj in data:
                dataset.append(
                    self.export_resource(obj), fields, num_indent=self.NUM_INDENT, color=color
                )
        else:
            dataset.append(
                self.export_resource(data), fields, num_indent=self.NUM_INDENT, color=color
            )

        return dataset


class BatchResource(MultiModelResource):
    class Meta:
        model = models.Batch
        include = ("label",)


class RecordResource(MultiModelResource):
    batch_id = fields.Field(attribute="batch_id")
    label = ExcelField(attribute="label", column_name="Defendent Name")
    dob = ExcelField(attribute="dob", column_name="Date of Birth")
    has_additional_offenses = ExcelField(
        attribute="has_additional_offenses", column_name="Additional Offenses Exist"
    )
    jurisdiction = ExcelField(
        attribute="jurisdiction",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.JURISDICTION_MAP.values())}"',
            showDropDown=True,
        ),
    )
    sex = ExcelField(
        attribute="sex",
        dropdown=DataValidation(
            type="list", formula1=f'"{",".join(constants.SEX_MAP.keys())}"', showDropDown=True
        ),
    )

    def dehydrate_jurisdiction(self, record):
        jurisdiction = getattr(record, "jurisdiction")
        return constants.JURISDICTION_MAP.get(jurisdiction, constants.NOT_AVAILABLE)

    def hydrate_jurisdiction(self, field, data):
        attribute = field.attribute
        hydrated_value = constants.NOT_AVAILABLE
        for k, v in constants.JURISDICTION_MAP.items():
            if v == data[attribute]:
                hydrated_value = k
                break
        data[attribute] = hydrated_value
        return data

    def dehydrate_sex(self, record):
        sex = getattr(record, "sex")
        try:
            sex_value = constants.SEX_CHOICES[sex]
        except:
            sex_value = constants.NOT_AVAILABLE
        return sex_value

    def hydrate_sex(self, field, data):
        attribute = field.attribute
        data[attribute] = constants.SEX_MAP.get(data[attribute], constants.NOT_AVAILABLE)
        return data

    def get_export_fields(self):
        return [
            field for field in super().get_export_fields() if field.column_name not in ("batch_id",)
        ]

    def get_export_headers(self):
        return [header for header in super().get_export_headers() if header not in ("Batch Id",)]

    class Meta:
        model = models.CIPRSRecord
        exclude = (
            "id",
            "batch",
            "batch_file",
            "date_uploaded",
            "data",
        )
        export_order = (
            "batch_id",
            "label",
            "file_no",
            "dob",
            "jurisdiction",
            "county",
            "has_additional_offenses",
        )
        force_init_instance = True


class OffenseResource(MultiModelResource):
    ciprs_record_id = fields.Field(attribute="ciprs_record_id")
    disposed_on = ExcelField(attribute="disposed_on", column_name="Disposition Date")
    jurisdiction = ExcelField(
        attribute="jurisdiction",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.JURISDICTION_MAP.values())}"',
            showDropDown=True,
        ),
    )
    plea = ExcelField(
        attribute="plea",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.VERDICT_CODE_MAP.keys())}"',
            showDropDown=True,
        ),
    )
    verdict = ExcelField(
        attribute="verdict",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.VERDICT_CODE_MAP.keys())}"',
            showDropDown=True,
        ),
    )
    disposition_method = ExcelField(
        attribute="disposition_method",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.DISPOSITION_METHOD_CODE_MAP.keys())}"',
            showDropDown=True,
        ),
    )

    def get_export_fields(self):
        return [
            field for field in super().get_export_fields() if field.column_name != "ciprs_record_id"
        ]

    def get_export_headers(self):
        return [header for header in super().get_export_headers() if header != "Ciprs Record Id"]

    def dehydrate_jurisdiction(self, record):
        jurisdiction = getattr(record, "jurisdiction")
        return constants.JURISDICTION_MAP[jurisdiction]

    def hydrate_jurisdiction(self, field, data):
        attribute = field.attribute
        for k, v in constants.JURISDICTION_MAP.items():
            if v == data[attribute]:
                data[attribute] = k
                break

        return data

    class Meta:
        model = models.Offense
        exclude = ("id", "ciprs_record")
        export_order = ("ciprs_record_id",)
        force_init_instance = True


class OffenseRecordResource(MultiModelResource):
    NUM_INDENT = 1

    offense_id = fields.Field(attribute="offense_id")
    count = ExcelField(attribute="count", column_name="#")
    severity = ExcelField(
        attribute="severity",
        dropdown=DataValidation(
            type="list",
            formula1=f'"{",".join(constants.SEVERITIES._db_values)}"',
            showDropDown=True,
        ),
    )
    action = ExcelField(
        attribute="action",
        dropdown=DataValidation(
            type="list", formula1=f'"{",".join(constants.ACTIONS._db_values)}"', showDropDown=True
        ),
    )

    def get_export_fields(self):
        return [field for field in super().get_export_fields() if field.column_name != "offense_id"]

    def get_export_headers(self):
        return [header for header in super().get_export_headers() if header not in ("Offense Id")]

    class Meta:
        model = models.OffenseRecord
        exclude = ("id", "agency", "offense")
        export_order = ("offense_id",)
        force_init_instance = True


class ClientResource(MultiModelResource):
    dob = ExcelField(attribute="dob", column_name="Date of Birth")

    class Meta:
        model = models.Client
        exclude = ("id", "contact_ptr", "category", "user")
        force_init_instance = True

    def get_import_fields(self):
        return ["batch_id"] + super().get_import_fields()


class RecordSummaryResource(resources.ModelResource):
    class Meta:
        model = models.Batch

    def __init__(self):
        super().__init__()
        self.batch_resource = BatchResource()
        self.client_resource = ClientResource()
        self.record_resource = RecordResource()
        self.record_headers = self.record_resource.get_export_headers()
        self.first_record_header = self.record_headers[0]
        self.offense_resource = OffenseResource()
        self.offense_headers = self.offense_resource.get_export_headers()
        self.first_offense_header = self.offense_headers[0]
        self.offense_record_resource = OffenseRecordResource()
        self.offense_record_headers = self.offense_record_resource.get_export_headers()
        self.first_offense_record_header = self.offense_record_headers[0]

    def export(self, batch_object=None, *args, **kwargs):
        if not isinstance(batch_object, models.Batch):
            raise ValueError("Queryset must be a Batch object")

        dataset = ExcelDataset()

        # Client
        if batch_object.client:
            dataset.create_new_worksheet(title=CLIENT_SHEET_TITLE)
            self.client_resource.export_excel(batch_object.client, dataset)

        # Record
        for record_object in (
            batch_object.records.order_by("file_no")
            .prefetch_related("offenses", "offenses__offense_records")
            .all()
        ):
            dataset.create_new_worksheet(title=record_object.file_no)
            self.record_resource.export_excel(record_object, dataset, color="E0FFFF")
            dataset.append_separator()

            # Offense
            i = 0
            for offense_object in record_object.offenses.all():
                color = "90EE90" if i % 2 == 1 else "FFFFE0"
                self.offense_resource.export_excel(offense_object, dataset, color=color)

                # Offense Record
                self.offense_record_resource.export_excel(
                    offense_object.offense_records.all(), dataset, color=color
                )
                dataset.append_separator()
                i += 1

        dataset.resize_columns()

        return dataset

    def import_data(self, workbook: Workbook, batch: models.Batch, *args, **kwargs):
        with transaction.atomic():
            if CLIENT_SHEET_TITLE in workbook:
                client_worksheet = workbook[CLIENT_SHEET_TITLE]
                client_headers = [cell.value for cell in client_worksheet[1]]
                client_values = [cell.value for cell in client_worksheet[2]]
                client_dataset = tablib.Dataset()
                client_dataset.headers = client_headers
                client_dataset.append(client_values)
                self.client_resource.import_data(client_dataset)
                workbook.remove(client_worksheet)

            current_resource = None
            current_dataset = tablib.Dataset()
            for worksheet in workbook.worksheets:
                if current_resource and current_dataset:
                    result = current_resource.import_data(current_dataset)
                current_resource = None
                current_dataset = tablib.Dataset()
                for row_number, row in enumerate(worksheet.iter_rows(), 1):
                    if all(cell.value in (None, "") for cell in row):
                        continue  # This is an empty row.

                    for cell_number, cell in enumerate(worksheet[row_number]):
                        if cell.value is not None:
                            break  # Get first non-null value

                    if cell.value in (
                        self.first_record_header,
                        self.first_offense_header,
                        self.first_offense_record_header,
                    ):
                        if current_resource and current_dataset:
                            result = current_resource.import_data(current_dataset)
                            current_dataset = tablib.Dataset()

                        if cell.value == self.first_record_header:
                            current_resource = self.record_resource
                            current_dataset.headers = [
                                col.column_name for col in current_resource.get_import_fields()
                            ]
                        elif cell.value == self.first_offense_header:
                            current_resource = self.offense_resource
                            current_dataset.headers = [
                                col.column_name for col in current_resource.get_import_fields()
                            ]
                        elif cell.value == self.first_offense_record_header:
                            current_resource = self.offense_record_resource
                            current_dataset.headers = [
                                col.column_name for col in current_resource.get_import_fields()
                            ]
                    else:
                        if current_resource == self.record_resource:
                            row = list(row)[
                                current_resource.NUM_INDENT : len(self.record_headers)
                                + current_resource.NUM_INDENT
                            ]
                            current_dataset.append([batch.id] + [cell.value for cell in row])
                        elif current_resource == self.offense_resource:
                            row = list(row)[
                                current_resource.NUM_INDENT : len(self.offense_headers)
                                + current_resource.NUM_INDENT
                            ]
                            record_id = self.record_resource.saved_instance_ids[-1]
                            current_dataset.append([record_id] + [cell.value for cell in row])
                        elif current_resource == self.offense_record_resource:
                            row = list(row)[
                                current_resource.NUM_INDENT : len(self.offense_record_headers)
                                + current_resource.NUM_INDENT
                            ]
                            offense_id = self.offense_resource.saved_instance_ids[-1]
                            current_dataset.append([offense_id] + [cell.value for cell in row])
                        else:
                            current_dataset.append(row[cell_number:])

            if current_resource and current_dataset:
                result = current_resource.import_data(current_dataset)

            batch.refresh_from_db()
            create_batch_petitions(batch)
